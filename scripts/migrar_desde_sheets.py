#!/usr/bin/env python
"""
Script de migración: importa datos desde "STOCKPLUS NG v7.xlsx" a PostgreSQL.

Uso:
    python scripts/migrar_desde_sheets.py           # migración real
    python scripts/migrar_desde_sheets.py --dry-run # solo muestra qué haría

Estructura real del xlsx (descubierta con inspección previa):
  - PRODUCTOS   : 323 filas, CLAVE_PRODUCTO es la clave única por variante
  - PRODUCTOS_WEB: vincula por CLAVE_PUBLICACION (= CLAVE_PRODUCTO en mayúsculas)
  - LOTES       : 7 lotes (LOT-00000x / LOTE-0000x)
  - COMPRAS     : 984 filas, precio en PRECIO_COMPRA_UNITARIO
  - VENTAS      : 991 filas, sin columna CANAL (solo ORIGEN_CARGA='WEB')
  - GASTOS      : 19 filas

Notas de diseño:
  - Produto.codigo (max_length=20) no puede contener CLAVE_PRODUTO (hasta 76 chars).
    Estrategia:
      · COD_PRODUTO con 1 variante  → codigo = COD_PRODUTO (ej: 'PROD-0503')
      · COD_PRODUTO con N variantes → codigo = COD_PRODUTO + '-' + contador 2 dígitos
        ej: 'PROD-0001-01', 'PROD-0001-38'  (max 14 chars)
  - Para COMPRAS/VENTAS, que referencian por COD_PRODUTO+MODELO+COLOR,
    se construye un variant_map antes de migrar.
"""

import os
import sys
import argparse
import traceback
from collections import Counter
from decimal import Decimal, InvalidOperation
from datetime import datetime, date

# ── Django setup ─────────────────────────────────────────────────────────────
sys.path.insert(0, '/home/sysadmin/stockplus')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')

import django
django.setup()

# ── Terceros ──────────────────────────────────────────────────────────────────
import openpyxl

# ── Modelos Django ────────────────────────────────────────────────────────────
from apps.inventario.models import Producto, StockActual
from apps.compras.models    import Compra, ItemCompra
from apps.ventas.models     import Venta, ItemVenta
from apps.ecommerce.models  import ProductoWeb
from apps.gastos.models     import ConceptoAdicional, GastoGeneral

# ── Constantes ────────────────────────────────────────────────────────────────
XLSX_PATH       = '/home/sysadmin/documentos e imagenes/STOCKPLUS NG v7.xlsx'
VENTA_MIG_PREFIX = 'MIGS'   # prefijo para Venta.numero en migración (idempotencia)


# ═════════════════════════════════════════════════════════════════════════════
# Helpers de conversión
# ═════════════════════════════════════════════════════════════════════════════

def to_decimal(value, default=Decimal('0')):
    if value is None or str(value).strip() == '':
        return default
    try:
        return Decimal(str(value)).quantize(Decimal('1'))
    except (InvalidOperation, ValueError):
        return default


def to_int(value, default=0):
    if value is None or str(value).strip() == '':
        return default
    try:
        return int(float(str(value)))
    except (ValueError, TypeError):
        return default


def to_str(value, default=''):
    if value is None:
        return default
    return str(value).strip()


def to_date(value):
    """Convierte cualquier formato de fecha a date. Retorna None si falla."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip().split('.')[0]
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y',
                '%d-%m-%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def bool_xlsx(value):
    """Convierte SI/NO/TRUE/FALSE/1/0 a bool."""
    return to_str(value).upper().strip() in ('SI', 'SÍ', 'YES', 'TRUE', '1', 'VERDADERO')


def normalizar_texto(texto):
    """
    Normaliza un texto para comparaciones:
    - minúsculas
    - elimina tildes y caracteres diacríticos (á→a, ñ→n, etc.)
    - elimina espacios extremos

    Necesario porque CLAVE_PRODUCTO en la hoja PRODUCTOS no tiene tildes
    (ej: 'mono') pero CLAVE_PUBLICACION en PRODUCTOS_WEB sí las tiene ('moño').
    """
    import unicodedata
    texto = to_str(texto).lower().strip()
    # Descompone caracteres compuestos y elimina marcas diacríticas
    nfkd = unicodedata.normalize('NFKD', texto)
    return ''.join(c for c in nfkd if not unicodedata.combining(c))


def mapear_canal(origen_carga):
    """
    VENTAS no tiene columna CANAL; solo tiene ORIGEN_CARGA.
    WEB = ingresado desde la app web → ecommerce.
    """
    mapping = {
        'WEB':        'ecommerce',
        'WHATSAPP':   'whatsapp',
        'WA':         'whatsapp',
        'INSTAGRAM':  'instagram',
        'IG':         'instagram',
        'FACEBOOK':   'facebook',
        'FB':         'facebook',
        'PRESENCIAL': 'presencial',
        'APP':        'otro',
    }
    return mapping.get(to_str(origen_carga).upper().strip(), 'otro')


def read_sheet(wb, sheet_name):
    """
    Lee una hoja y devuelve lista de dicts usando la primera fila como cabeceras.
    Omite filas completamente vacías.
    """
    if sheet_name not in wb.sheetnames:
        print(f"  [WARN] Hoja '{sheet_name}' no encontrada en el archivo")
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [to_str(h) for h in rows[0]]
    result = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        result.append(dict(zip(headers, row)))
    return result


# ═════════════════════════════════════════════════════════════════════════════
# Generación de códigos únicos para productos con variantes de color/modelo
# ═════════════════════════════════════════════════════════════════════════════

def construir_mapas_productos(rows_productos):
    """
    Construye dos diccionarios para manejar variantes de producto:

    1. clave_a_codigo:  {clave_producto_lower: codigo_unico}
       Clave única por variante. Permite buscar desde PRODUCTOS_WEB.

    2. coords_a_codigo: {(cod_upper, modelo_upper, color_upper): codigo_unico}
       Permite buscar desde COMPRAS/VENTAS que solo dan COD_PRODUTO+MODELO+COLOR.

    Estrategia de código:
      · Un único COD_PRODUTO  → codigo = COD_PRODUTO         (ej: 'PROD-0503')
      · COD_PRODUTO repetido  → codigo = COD_PRODUTO-NN       (ej: 'PROD-0001-01')
    """
    cnt_por_cod = Counter(
        to_str(r.get('COD_PRODUCTO')).upper()
        for r in rows_productos
        if r.get('COD_PRODUCTO')
    )

    variante_idx = {}   # cod_upper → counter actual
    clave_a_codigo  = {}
    coords_a_codigo = {}

    for r in rows_productos:
        cod   = to_str(r.get('COD_PRODUCTO')).upper().strip()
        clave = normalizar_texto(r.get('CLAVE_PRODUCTO') or '')
        modelo= to_str(r.get('MODELO')).upper().strip()
        color = to_str(r.get('COLOR')).upper().strip()

        if not cod or not clave:
            continue

        if cnt_por_cod[cod] == 1:
            codigo = cod  # Ej: 'PROD-0503'
        else:
            variante_idx[cod] = variante_idx.get(cod, 0) + 1
            codigo = f"{cod}-{variante_idx[cod]:02d}"  # Ej: 'PROD-0001-01'

        # Usar clave normalizada (sin tildes) como key para tolerar diferencias
        # entre hojas (ej: PRODUCTOS usa 'mono', PRODUCTOS_WEB usa 'moño')
        clave_norm = normalizar_texto(clave)
        clave_a_codigo[clave_norm]               = codigo
        coords_a_codigo[(cod, modelo, color)]    = codigo

    return clave_a_codigo, coords_a_codigo


def buscar_codigo(cod_producto, modelo, color, coords_a_codigo):
    """
    Busca el codigo generado para un producto a partir de COD_PRODUTO+MODELO+COLOR.
    Si no hay coincidencia exacta, intenta con solo COD_PRODUTO (si es único).
    """
    key = (
        to_str(cod_producto).upper().strip(),
        to_str(modelo).upper().strip(),
        to_str(color).upper().strip(),
    )
    codigo = coords_a_codigo.get(key)
    if codigo:
        return codigo

    # Fallback: si solo hay una variante para ese COD_PRODUTO, devuelve esa
    cod_upper = to_str(cod_producto).upper().strip()
    matches = [v for (c, m, col), v in coords_a_codigo.items() if c == cod_upper]
    if len(set(matches)) == 1:
        return matches[0]

    return None


# ═════════════════════════════════════════════════════════════════════════════
# Clase de contadores / resumen
# ═════════════════════════════════════════════════════════════════════════════

class Contadores:
    def __init__(self):
        self.productos_creados       = 0
        self.productos_omitidos      = 0
        self.stock_actualizado       = 0
        self.productos_web_creados   = 0
        self.productos_web_omitidos  = 0
        self.lotes_creados           = 0
        self.lotes_omitidos          = 0
        self.items_compra_creados    = 0
        self.items_compra_omitidos   = 0
        self.ventas_creadas          = 0
        self.ventas_omitidas         = 0
        self.gastos_creados          = 0
        self.gastos_omitidos         = 0
        self.warnings                = []

    def warn(self, msg):
        self.warnings.append(msg)
        print(f"    [WARN] {msg}")

    def resumen(self):
        sep = '=' * 64
        print(f'\n{sep}')
        print('  RESUMEN DE MIGRACION')
        print(sep)
        print(f'  Productos creados       : {self.productos_creados}')
        print(f'  Productos ya existentes : {self.productos_omitidos}')
        print(f'  Stock actualizado       : {self.stock_actualizado}')
        print(f'  ProductosWeb creados    : {self.productos_web_creados}')
        print(f'  ProductosWeb existentes : {self.productos_web_omitidos}')
        print(f'  Lotes (Compras) creados : {self.lotes_creados}')
        print(f'  Lotes ya existentes     : {self.lotes_omitidos}')
        print(f'  Items de compra creados : {self.items_compra_creados}')
        print(f'  Items compra existentes : {self.items_compra_omitidos}')
        print(f'  Ventas creadas          : {self.ventas_creadas}')
        print(f'  Ventas omitidas         : {self.ventas_omitidas}')
        print(f'  Gastos creados          : {self.gastos_creados}')
        print(f'  Gastos omitidos         : {self.gastos_omitidos}')
        if self.warnings:
            print(f'\n  [!] {len(self.warnings)} advertencias (ver arriba)')
        print(sep)


# ═════════════════════════════════════════════════════════════════════════════
# Paso 1: PRODUCTOS  →  Producto + StockActual
# ═════════════════════════════════════════════════════════════════════════════

def migrar_productos(rows, clave_a_codigo, dry_run, cnt):
    total = len(rows)
    label = '[DRY-RUN] ' if dry_run else ''
    print(f'\n{label}[1/6] Migrando productos (PRODUCTOS)...')

    for i, row in enumerate(rows, 1):
        clave     = normalizar_texto(row.get('CLAVE_PRODUCTO') or '')
        cod_base  = to_str(row.get('COD_PRODUCTO')).upper().strip()
        codigo    = clave_a_codigo.get(clave)

        if not codigo:
            cnt.warn(f'PRODUCTOS fila {i}: sin CLAVE_PRODUCTO o COD_PRODUCTO, se omite')
            cnt.productos_omitidos += 1
            continue

        nombre      = (to_str(row.get('NOMBRE_PRODUCTO'))
                       or to_str(row.get('DESCRIPCION_PRODUCTO'))
                       or codigo)
        modelo      = to_str(row.get('MODELO'))
        color       = to_str(row.get('COLOR'))
        descripcion = to_str(row.get('DESCRIPCION_PRODUCTO'))
        precio_costo= to_decimal(row.get('PRECIO_COMPRA_REF'))
        precio_venta= to_decimal(row.get('PRECIO_VENTA_AUTO'))
        stock_disp  = to_int(row.get('STOCK_DISPONIBLE'))
        activo      = bool_xlsx(row.get('ACTIVO') or 'SI')

        if dry_run:
            print(f'  [{i:>4}/{total}] DRY Producto {codigo!r:16} '
                  f'{nombre[:35]!r:37} stock={stock_disp:>3} '
                  f'costo={precio_costo:>10,} venta={precio_venta:>10,}')
            cnt.productos_creados += 1
            continue

        try:
            prod, created = Producto.objects.get_or_create(
                codigo=codigo,
                defaults={
                    'nombre':         nombre,
                    'modelo_celular': modelo,
                    'color':          color,
                    'descripcion':    descripcion,
                    'precio_costo':   precio_costo,
                    'precio_venta':   precio_venta,
                    'activo':         activo,
                }
            )
            if created:
                cnt.productos_creados += 1
            else:
                cnt.productos_omitidos += 1

            # Crear o actualizar StockActual
            stock_obj, s_created = StockActual.objects.get_or_create(producto=prod)
            if s_created or stock_obj.cantidad != stock_disp:
                stock_obj.cantidad = stock_disp
                stock_obj.save(update_fields=['cantidad', 'actualizado_en'])
                if not s_created:
                    cnt.stock_actualizado += 1

        except Exception as exc:
            cnt.warn(f'PRODUCTOS fila {i} ({codigo}): {exc}')
            cnt.productos_omitidos += 1

        if i % 50 == 0 or i == total:
            print(f'  Productos procesados: {i}/{total}', end='\r')

    print(f'  Migrando productos... {total}/{total} OK')


# ═════════════════════════════════════════════════════════════════════════════
# Paso 2: PRODUCTOS_WEB  →  ProductoWeb
# Vincula por CLAVE_PUBLICACION (= CLAVE_PRODUTO en mayúsculas)
# ═════════════════════════════════════════════════════════════════════════════

def migrar_productos_web(rows, clave_a_codigo, dry_run, cnt):
    total = len(rows)
    label = '[DRY-RUN] ' if dry_run else ''
    print(f'\n{label}[2/6] Migrando productos web (PRODUCTOS_WEB)...')

    # Cache de productos en DB (no cargar en dry-run)
    prod_cache = {} if dry_run else {
        p.codigo: p for p in Producto.objects.all()
    }

    for i, row in enumerate(rows, 1):
        clave_pub = normalizar_texto(row.get('CLAVE_PUBLICACION') or '')
        if not clave_pub:
            cnt.warn(f'PRODUCTOS_WEB fila {i}: CLAVE_PUBLICACION vacia, se omite')
            cnt.productos_web_omitidos += 1
            continue

        codigo = clave_a_codigo.get(clave_pub)
        if not codigo:
            cnt.warn(f'PRODUCTOS_WEB fila {i}: CLAVE_PUBLICACION {clave_pub!r:.40} '
                     f'no mapeada, se omite')
            cnt.productos_web_omitidos += 1
            continue

        titulo_web      = to_str(row.get('TITULO_WEB'))
        descripcion_web = to_str(row.get('DESCRIPCION_WEB'))
        precio_web      = to_decimal(row.get('PRECIO_WEB'))
        visible         = bool_xlsx(row.get('VISIBLE'))
        destacado       = bool_xlsx(row.get('DESTACADO'))
        imagen_url      = to_str(row.get('IMAGE_URL'))

        if dry_run:
            print(f'  [{i:>4}/{total}] DRY ProductoWeb {codigo!r:16} '
                  f'visible={visible} destacado={destacado} precio={precio_web:>10,}')
            cnt.productos_web_creados += 1
            continue

        prod = prod_cache.get(codigo)
        if prod is None:
            cnt.warn(f'PRODUCTOS_WEB fila {i}: Producto {codigo!r} no existe en DB, se omite')
            cnt.productos_web_omitidos += 1
            continue

        try:
            pw, created = ProductoWeb.objects.get_or_create(
                producto=prod,
                defaults={
                    'titulo_web':      titulo_web,
                    'descripcion_web': descripcion_web,
                    'precio_web':      precio_web,
                    'visible':         visible,
                    'destacado':       destacado,
                }
            )
            if created:
                cnt.productos_web_creados += 1
            else:
                cnt.productos_web_omitidos += 1

            # imagen_url vive en Producto (ProductoWeb no tiene ese campo)
            if imagen_url and not prod.imagen_url:
                prod.imagen_url = imagen_url
                prod.save(update_fields=['imagen_url'])

        except Exception as exc:
            cnt.warn(f'PRODUCTOS_WEB fila {i} ({codigo}): {exc}')
            cnt.productos_web_omitidos += 1

        if i % 50 == 0 or i == total:
            print(f'  ProductosWeb procesados: {i}/{total}', end='\r')

    print(f'  Migrando productos web... {total}/{total} OK')


# ═════════════════════════════════════════════════════════════════════════════
# Paso 3: LOTES  →  Compra (una Compra por COD_LOTE)
# ═════════════════════════════════════════════════════════════════════════════

def migrar_lotes(rows, dry_run, cnt):
    total = len(rows)
    label = '[DRY-RUN] ' if dry_run else ''
    print(f'\n{label}[3/6] Migrando lotes (LOTES -> Compra)...')

    for i, row in enumerate(rows, 1):
        cod_lote = to_str(row.get('COD_LOTE')).strip()
        if not cod_lote:
            cnt.warn(f'LOTES fila {i}: COD_LOTE vacio, se omite')
            cnt.lotes_omitidos += 1
            continue

        fecha_carga   = to_date(row.get('FECHA_CARGA')) or date.today()
        proveedor_ref = to_str(row.get('PROVEEDOR_REFERENCIA'))
        obs_parts     = [x for x in [
            f'Proveedor: {proveedor_ref}' if proveedor_ref else '',
            to_str(row.get('OBSERVACIONES')),
        ] if x]
        obs = ' | '.join(obs_parts)

        if dry_run:
            print(f'  [{i:>4}/{total}] DRY Compra {cod_lote!r:14} '
                  f'fecha={fecha_carga} proveedor={proveedor_ref!r}')
            cnt.lotes_creados += 1
            continue

        try:
            compra, created = Compra.objects.get_or_create(
                numero=cod_lote,
                defaults={
                    'fecha':         fecha_carga,
                    'estado':        'recibida',
                    'observaciones': obs,
                }
            )
            if created:
                cnt.lotes_creados += 1
            else:
                cnt.lotes_omitidos += 1

        except Exception as exc:
            cnt.warn(f'LOTES fila {i} ({cod_lote}): {exc}')
            cnt.lotes_omitidos += 1

        if i % 100 == 0 or i == total:
            print(f'  Lotes procesados: {i}/{total}', end='\r')

    print(f'  Migrando lotes... {total}/{total} OK')


# ═════════════════════════════════════════════════════════════════════════════
# Paso 4: COMPRAS  →  ItemCompra
# ═════════════════════════════════════════════════════════════════════════════

def migrar_compras(rows, coords_a_codigo, dry_run, cnt):
    total = len(rows)
    label = '[DRY-RUN] ' if dry_run else ''
    print(f'\n{label}[4/6] Migrando items de compra (COMPRAS -> ItemCompra)...')

    # Caches para evitar consultas repetidas
    _compras  = {}
    _productos = {}

    for i, row in enumerate(rows, 1):
        cod_lote  = to_str(row.get('COD_LOTE')).strip()
        cod_raw   = to_str(row.get('COD_PRODUCTO')).upper().strip()
        modelo    = to_str(row.get('MODELO')).upper().strip()
        color     = to_str(row.get('COLOR')).upper().strip()
        cantidad  = to_int(row.get('CANTIDAD'), default=1)
        precio    = to_decimal(row.get('PRECIO_COMPRA_UNITARIO'))

        if not cod_lote or not cod_raw:
            cnt.warn(f'COMPRAS fila {i}: COD_LOTE o COD_PRODUCTO vacio, se omite')
            cnt.items_compra_omitidos += 1
            continue

        codigo = buscar_codigo(cod_raw, modelo, color, coords_a_codigo)
        if not codigo:
            cnt.warn(f'COMPRAS fila {i}: no se encontro variante para '
                     f'{cod_raw}/{modelo}/{color}, se omite')
            cnt.items_compra_omitidos += 1
            continue

        if dry_run:
            print(f'  [{i:>4}/{total}] DRY ItemCompra lote={cod_lote!r:12} '
                  f'prod={codigo!r:16} cant={cantidad} precio={precio:>10,}')
            cnt.items_compra_creados += 1
            continue

        # Buscar Compra (cache)
        if cod_lote not in _compras:
            try:
                _compras[cod_lote] = Compra.objects.get(numero=cod_lote)
            except Compra.DoesNotExist:
                cnt.warn(f'COMPRAS fila {i}: Lote {cod_lote!r} no existe en DB, se omite')
                _compras[cod_lote] = None
        compra = _compras[cod_lote]
        if compra is None:
            cnt.items_compra_omitidos += 1
            continue

        # Buscar Producto (cache)
        if codigo not in _productos:
            try:
                _productos[codigo] = Producto.objects.get(codigo=codigo)
            except Producto.DoesNotExist:
                cnt.warn(f'COMPRAS fila {i}: Producto {codigo!r} no existe en DB, se omite')
                _productos[codigo] = None
        prod = _productos[codigo]
        if prod is None:
            cnt.items_compra_omitidos += 1
            continue

        # Si precio es 0, usar precio_costo del producto
        if precio == 0:
            precio = prod.precio_costo

        try:
            item, created = ItemCompra.objects.get_or_create(
                compra=compra,
                producto=prod,
                defaults={
                    'cantidad':        max(cantidad, 1),
                    'precio_unitario': precio,
                }
            )
            if created:
                cnt.items_compra_creados += 1
            else:
                cnt.items_compra_omitidos += 1

        except Exception as exc:
            cnt.warn(f'COMPRAS fila {i} ({cod_lote}/{codigo}): {exc}')
            cnt.items_compra_omitidos += 1

        if i % 100 == 0 or i == total:
            print(f'  Items de compra procesados: {i}/{total}', end='\r')

    print(f'  Migrando items de compra... {total}/{total} OK')


# ═════════════════════════════════════════════════════════════════════════════
# Paso 5: VENTAS  →  Venta + ItemVenta
#
# Nota: VENTAS no tiene columna CANAL; solo ORIGEN_CARGA ('WEB' = ecommerce).
# Idempotencia: numero = MIGS-{i:06d} → get_or_create no duplica en 2da ejecución.
# ═════════════════════════════════════════════════════════════════════════════

def migrar_ventas(rows, coords_a_codigo, dry_run, cnt):
    total = len(rows)
    label = '[DRY-RUN] ' if dry_run else ''
    print(f'\n{label}[5/6] Migrando ventas (VENTAS -> Venta + ItemVenta)...')

    _productos = {}

    for i, row in enumerate(rows, 1):
        fecha       = to_date(row.get('FECHA'))
        cod_raw     = to_str(row.get('COD_PRODUCTO')).upper().strip()
        modelo      = to_str(row.get('MODELO')).upper().strip()
        color       = to_str(row.get('COLOR')).upper().strip()
        cod_lote    = to_str(row.get('COD_LOTE')).strip()
        cantidad    = to_int(row.get('CANTIDAD'), default=1)
        precio_vta  = to_decimal(row.get('PRECIO_VENTA_REAL'))
        canal       = mapear_canal(row.get('ORIGEN_CARGA'))
        cliente_nom = to_str(row.get('CLIENTE'))
        numero_mig  = f'{VENTA_MIG_PREFIX}-{i:06d}'

        if fecha is None:
            cnt.warn(f'VENTAS fila {i}: FECHA invalida, se omite')
            cnt.ventas_omitidas += 1
            continue
        if not cod_raw:
            cnt.warn(f'VENTAS fila {i}: COD_PRODUCTO vacio, se omite')
            cnt.ventas_omitidas += 1
            continue

        codigo = buscar_codigo(cod_raw, modelo, color, coords_a_codigo)
        if not codigo:
            cnt.warn(f'VENTAS fila {i}: no se encontro variante para '
                     f'{cod_raw}/{modelo}/{color}, se omite')
            cnt.ventas_omitidas += 1
            continue

        if dry_run:
            print(f'  [{i:>4}/{total}] DRY Venta {numero_mig} {fecha} '
                  f'{codigo!r:16} cant={cantidad} precio={precio_vta:>10,} canal={canal}')
            cnt.ventas_creadas += 1
            continue

        # Buscar Producto (cache)
        if codigo not in _productos:
            try:
                _productos[codigo] = Producto.objects.get(codigo=codigo)
            except Producto.DoesNotExist:
                cnt.warn(f'VENTAS fila {i}: Producto {codigo!r} no existe en DB, se omite')
                _productos[codigo] = None
        prod = _productos[codigo]
        if prod is None:
            cnt.ventas_omitidas += 1
            continue

        obs_parts = [x for x in [
            f'Lote: {cod_lote}' if cod_lote else '',
            to_str(row.get('OBSERVACIONES')),
        ] if x]

        try:
            venta, v_created = Venta.objects.get_or_create(
                numero=numero_mig,
                defaults={
                    'fecha':          fecha,
                    'estado':         'entregada',
                    'canal':          canal,
                    'cliente_nombre': cliente_nom,
                    'observaciones':  ' | '.join(obs_parts),
                }
            )

            if v_created:
                ItemVenta.objects.create(
                    venta=venta,
                    producto=prod,
                    cantidad=max(cantidad, 1),
                    precio_unitario=precio_vta,
                    costo_unitario=prod.precio_costo,
                )
                cnt.ventas_creadas += 1
            else:
                cnt.ventas_omitidas += 1

        except Exception as exc:
            cnt.warn(f'VENTAS fila {i} ({codigo}): {exc}')
            traceback.print_exc()
            cnt.ventas_omitidas += 1

        if i % 100 == 0 or i == total:
            print(f'  Ventas procesadas: {i}/{total}', end='\r')

    print(f'  Migrando ventas... {total}/{total} OK')


# ═════════════════════════════════════════════════════════════════════════════
# Paso 6: GASTOS  →  GastoGeneral
# ═════════════════════════════════════════════════════════════════════════════

def migrar_gastos(rows, dry_run, cnt):
    total = len(rows)
    label = '[DRY-RUN] ' if dry_run else ''
    print(f'\n{label}[6/6] Migrando gastos (GASTOS -> GastoGeneral)...')

    concepto_otro = None
    if not dry_run:
        concepto_otro, _ = ConceptoAdicional.objects.get_or_create(
            nombre='Otro gasto general',
            defaults={
                'tipo':        'egreso',
                'aplica_a':    'gasto',
                'descripcion': 'Gasto general sin categoria especifica',
            }
        )

    _conceptos = {}  # cache nombre_tipo → ConceptoAdicional

    for i, row in enumerate(rows, 1):
        fecha       = to_date(row.get('FECHA'))
        tipo_raw    = to_str(row.get('TIPO'))
        descripcion = to_str(row.get('DESCRIPCION'))
        monto       = to_decimal(row.get('MONTO'))
        proveedor   = to_str(row.get('PROVEEDOR'))

        if fecha is None:
            cnt.warn(f'GASTOS fila {i}: FECHA invalida, se omite')
            cnt.gastos_omitidos += 1
            continue

        desc_completa = ' | '.join(filter(None, [descripcion, proveedor]))

        if dry_run:
            print(f'  [{i:>4}/{total}] DRY Gasto {fecha} tipo={tipo_raw!r:20} '
                  f'monto={monto:>10,} desc={desc_completa!r}')
            cnt.gastos_creados += 1
            continue

        # Buscar o crear ConceptoAdicional para este TIPO
        if tipo_raw not in _conceptos:
            if tipo_raw:
                concepto = ConceptoAdicional.objects.filter(
                    nombre__icontains=tipo_raw
                ).first()
                if not concepto:
                    concepto, _ = ConceptoAdicional.objects.get_or_create(
                        nombre=tipo_raw,
                        defaults={
                            'tipo':     'egreso',
                            'aplica_a': 'gasto',
                        }
                    )
            else:
                concepto = concepto_otro
            _conceptos[tipo_raw] = concepto
        concepto = _conceptos.get(tipo_raw) or concepto_otro

        try:
            gasto, created = GastoGeneral.objects.get_or_create(
                fecha=fecha,
                concepto=concepto,
                monto=monto,
                defaults={
                    'descripcion': desc_completa or tipo_raw or 'Sin descripcion',
                }
            )
            if created:
                cnt.gastos_creados += 1
            else:
                cnt.gastos_omitidos += 1

        except Exception as exc:
            cnt.warn(f'GASTOS fila {i}: {exc}')
            cnt.gastos_omitidos += 1

    print(f'  Migrando gastos... {total}/{total} OK')


# ═════════════════════════════════════════════════════════════════════════════
# Verificación de integridad post-migración
# ═════════════════════════════════════════════════════════════════════════════

def verificar_integridad(wb, clave_a_codigo):
    sep = '=' * 64
    print(f'\n{sep}')
    print('  VERIFICACION DE INTEGRIDAD')
    print(sep)

    # -- Productos --
    rows_prod   = read_sheet(wb, 'PRODUCTOS')
    filas_prod  = sum(1 for r in rows_prod if r.get('CLAVE_PRODUCTO'))
    db_prods    = Producto.objects.count()
    estado_prod = 'OK' if db_prods >= filas_prod else f'FALTAN {filas_prod - db_prods}'
    print(f'  Productos  xlsx={filas_prod:>5}  DB={db_prods:>5}  [{estado_prod}]')

    # -- Ventas (items) --
    rows_ventas  = read_sheet(wb, 'VENTAS')
    filas_ventas = sum(1 for r in rows_ventas if r.get('COD_PRODUCTO'))
    db_items_v   = ItemVenta.objects.count()
    estado_v     = 'OK' if db_items_v >= filas_ventas else f'FALTAN {filas_ventas - db_items_v}'
    print(f'  ItemVentas xlsx={filas_ventas:>5}  DB={db_items_v:>5}  [{estado_v}]')

    # -- Lotes (Compras) --
    rows_lotes  = read_sheet(wb, 'LOTES')
    filas_lotes = sum(1 for r in rows_lotes if r.get('COD_LOTE'))
    db_compras  = Compra.objects.count()
    estado_c    = 'OK' if db_compras >= filas_lotes else f'FALTAN {filas_lotes - db_compras}'
    print(f'  Compras    xlsx={filas_lotes:>5}  DB={db_compras:>5}  [{estado_c}]')

    # -- Gastos --
    rows_gastos  = read_sheet(wb, 'GASTOS')
    filas_gastos = sum(1 for r in rows_gastos if r.get('FECHA'))
    db_gastos    = GastoGeneral.objects.count()
    estado_g     = 'OK' if db_gastos >= filas_gastos else f'FALTAN {filas_gastos - db_gastos}'
    print(f'  Gastos     xlsx={filas_gastos:>5}  DB={db_gastos:>5}  [{estado_g}]')

    # -- Coincidencia de stock por producto --
    print('\n  -- Verificacion de stock por producto --')
    diferencias = 0
    for row in rows_prod:
        clave = normalizar_texto(row.get('CLAVE_PRODUCTO') or '')
        codigo = clave_a_codigo.get(clave)
        if not codigo:
            continue
        stock_xlsx = to_int(row.get('STOCK_DISPONIBLE'))
        try:
            prod     = Producto.objects.get(codigo=codigo)
            stock_db = prod.stock.cantidad
            if stock_db != stock_xlsx:
                diferencias += 1
                if diferencias <= 5:
                    print(f'    {codigo}: xlsx={stock_xlsx}  DB={stock_db}')
        except Producto.DoesNotExist:
            diferencias += 1

    if diferencias == 0:
        print('  Todos los stocks coinciden  [OK]')
    else:
        print(f'  {diferencias} producto(s) con stock diferente entre xlsx y DB')

    print(sep)


# ═════════════════════════════════════════════════════════════════════════════
# Punto de entrada
# ═════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description='Migracion StockPlus NG v7.xlsx -> PostgreSQL'
    )
    parser.add_argument(
        '--dry-run', action='store_true',
        help='Solo muestra que haria sin escribir nada en la base de datos'
    )
    args = parser.parse_args()
    dry_run = args.dry_run

    if dry_run:
        print('\n[DRY-RUN] Modo simulacion: NO se escribira nada en la base de datos\n')
    else:
        print('\nIniciando migracion real...\n')

    print(f'  Archivo: {XLSX_PATH}')
    try:
        wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
    except FileNotFoundError:
        print(f'  ERROR: No se encontro el archivo: {XLSX_PATH}')
        sys.exit(1)
    except Exception as exc:
        print(f'  ERROR al abrir el archivo: {exc}')
        sys.exit(1)

    print(f'  Hojas disponibles: {wb.sheetnames}')

    # Leer hojas
    rows_productos     = read_sheet(wb, 'PRODUCTOS')
    rows_productos_web = read_sheet(wb, 'PRODUCTOS_WEB')
    rows_lotes         = read_sheet(wb, 'LOTES')
    rows_compras       = read_sheet(wb, 'COMPRAS')
    rows_ventas        = read_sheet(wb, 'VENTAS')
    rows_gastos        = read_sheet(wb, 'GASTOS')

    print(f'\n  Filas leidas:')
    print(f'    PRODUCTOS:     {len(rows_productos)}')
    print(f'    PRODUCTOS_WEB: {len(rows_productos_web)}')
    print(f'    LOTES:         {len(rows_lotes)}')
    print(f'    COMPRAS:       {len(rows_compras)}')
    print(f'    VENTAS:        {len(rows_ventas)}')
    print(f'    GASTOS:        {len(rows_gastos)}')

    # Construir mapas de variantes
    clave_a_codigo, coords_a_codigo = construir_mapas_productos(rows_productos)
    print(f'\n  Codigos de variantes generados: {len(clave_a_codigo)}')

    cnt = Contadores()

    # Ejecutar pasos en orden
    migrar_productos(rows_productos, clave_a_codigo, dry_run, cnt)
    migrar_productos_web(rows_productos_web, clave_a_codigo, dry_run, cnt)
    migrar_lotes(rows_lotes, dry_run, cnt)
    migrar_compras(rows_compras, coords_a_codigo, dry_run, cnt)
    migrar_ventas(rows_ventas, coords_a_codigo, dry_run, cnt)
    migrar_gastos(rows_gastos, dry_run, cnt)

    cnt.resumen()

    if not dry_run:
        verificar_integridad(wb, clave_a_codigo)

    wb.close()
    print('\nMigracion completada.\n')


if __name__ == '__main__':
    main()
